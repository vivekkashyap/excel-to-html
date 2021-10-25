from openpyxl import Workbook
from openpyxl import load_workbook

print('Task Started')
print('************************')

wb = load_workbook("Excel.xlsx")
sheet = wb['Sheet1']
r = sheet.max_row
print('Total Rows in Excel: ' + str(r))

filename = 'index.html'
try:
    f = open(filename, "x")
    # from which row you want to start the execution
    count = 0
    for i in range(2, r+1):
        name = sheet.cell(row=i, column=1).value;
        locationJoined = sheet.cell(row=i, column=2).value;
        team = sheet.cell(row=i, column=3).value;
        nativeLocation = sheet.cell(row=i, column=4).value;
        hobbies = sheet.cell(row=i, column=5).value;
        linkedinUrl = sheet.cell(row=i, column=6).value;
        photo = sheet.cell(row=i, column=7).value;
        emailId = sheet.cell(row=i, column=8).value;
        
        f.write('<html>\n')
        f.write('<body>\n')
        f.write('    <table>\n')
        f.write('        <!-- Person Name -->\n')
        f.write('        <h3>' + name +'</h3>\n')
        f.write('        <tr>\n')
        f.write('            <td>\n')
        f.write('                <!-- All Details -->\n')
        f.write('                <ul>\n')
        f.write('                    <li>Email: ' + emailId + '</li>\n')
        f.write('                    <li>Location: ' + locationJoined + '</li>\n')
        f.write('                    <li>Team: ' + team + '</li>\n')
        f.write('                    <li>Native Location: ' + nativeLocation + '</li>\n')
        f.write('                    <li>Hobbies: ' + hobbies + '</li>\n')
        f.write('                    <li>LinkedIn URL: ' + linkedinUrl + '</li>\n')
        f.write('                </ul> \n')
        f.write('            </td>\n')
        f.write('            <td>\n')
        f.write('                <!-- photo -->\n')
        f.write('                <img src="'+  photo + '" alt="Photo">\n')
        f.write('            </td>\n')
        f.write('        </tr>\n')
        f.write('    </table>\n')
        f.write('</body>\n')
        f.write('</html>\n')

        count = count + 1
        
    wb.save("Excel.xlsx")
    f.close()

    print('Entries Generated: ' + str(count))
    print('************************')
    print('Task Completed')
except IOError:
    print ('File ', filename, ' already exist. Please delete it!! OR Close the opened Excel File if opened!')
